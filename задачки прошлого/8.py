import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# === Чтение основной таблицы ===
df = pd.read_excel("production_res_valid_only copy.xlsx", skiprows=2, engine="openpyxl")

# Переименование
df = df.rename(columns={
    'Предприятие': 'Полное наименование',
    'ИНН': 'ИНН',
    'Дата внесения в реестр': 'Дата выдачи',
    'Срок действия': 'Дата окончания действия',
    'ОКПД2': 'ОКПД2'
})

# Только нужные
df = df[['Полное наименование', 'ИНН', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']].copy()

# Пустые поля
df['Компания'] = ''
df['Сфера'] = ''
df['Расшифровка ОКПД2'] = ''
df['Количество российской продукции, шт.'] = ''
df['Статус'] = ''

# Порядок колонок
df = df[['Компания', 'Полное наименование', 'ИНН', 'Дата выдачи',
         'Дата окончания действия', 'Сфера', 'ОКПД2', 'Расшифровка ОКПД2',
         'Количество российской продукции, шт.', 'Статус']]

# === Загрузка справочников ===
corresp_file = "gisp_соответствия.xlsx"
df_comp_name = pd.read_excel(corresp_file, sheet_name="Компания_Наименование")
df_sfera_okpd2 = pd.read_excel(corresp_file, sheet_name="Сфера_ОКПД2")
df_status = pd.read_excel(corresp_file, sheet_name="Наименование_Статус")

# === Функция для нормализации текста (убрать пробелы, \n, \xa0 и т.п.) ===
def clean(val):
    if pd.isna(val):
        return ''
    return str(val).strip().replace('\xa0', ' ').replace('\n', ' ').lower()

# === Создание словарей соответствия с нормализованными ключами ===
company_map = {clean(k): v for k, v in df_comp_name.values}
status_map = {clean(k): v for k, v in df_status.values}
sfera_map = {clean(k): v for k, v in df_sfera_okpd2.values}

# === Применение маппинга с очисткой ключей ===
df['Компания'] = df['Полное наименование'].apply(lambda x: company_map.get(clean(x), ''))
df['Статус'] = df['Полное наименование'].apply(lambda x: status_map.get(clean(x), ''))
df['Сфера'] = df['ОКПД2'].apply(lambda x: sfera_map.get(clean(x), ''))

# === Автозаполнение остальных строк с тем же полным названием ===
grouped = df.groupby("Полное наименование", sort=False)
for name, group in grouped:
    if len(group) > 1:
        first_row = group.iloc[0]
        for idx in group.index[1:]:
            for col in df.columns:
                if col not in ['ОКПД2', 'Расшифровка ОКПД2', 'Количество российской продукции, шт.']:
                    if pd.isna(df.at[idx, col]) or df.at[idx, col] == '':
                        df.at[idx, col] = first_row[col]

# === Запись в Excel ===
wb = Workbook()
ws = wb.active
ws.title = "Компании"

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Выравнивание
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", horizontal="left")

# Объединение одинаковых подряд ячеек
def merge_identical_cells(ws, col_letter):
    start_row = 2
    while start_row <= ws.max_row:
        current_value = ws[f"{col_letter}{start_row}"].value
        end_row = start_row
        while end_row + 1 <= ws.max_row and ws[f"{col_letter}{end_row + 1}"].value == current_value:
            end_row += 1
        if end_row > start_row:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
        start_row = end_row + 1

merge_identical_cells(ws, "B")  # Полное наименование
merge_identical_cells(ws, "C")  # ИНН

# Сохраняем
output_file = "сводка_по_компаниям.xlsx"
wb.save(output_file)

print(f"✅ Готово! Файл сохранён: {output_file}")
