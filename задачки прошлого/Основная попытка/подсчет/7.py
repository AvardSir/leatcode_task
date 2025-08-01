import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# === ПУТИ ===
summary_file = "сводка_по_компаниям.xlsx"
mapping_file = "gisp_соответствия.xlsx"

# === Читаем данные исходника, агрегируем ===
df = pd.read_excel("production_res_valid_only.xlsx", skiprows=2, engine="openpyxl")
df = df.rename(columns={
    'Предприятие': 'Полное наименование',
    'ИНН': 'ИНН',
    'Дата внесения в реестр': 'Дата выдачи',
    'Срок действия': 'Дата окончания действия',
    'ОКПД2': 'ОКПД2'
})
df = df[['Полное наименование', 'ИНН', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']].copy()
grouped = df.groupby(['ИНН', 'Полное наименование', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']) \
            .size().reset_index(name='Количество российской продукции, шт.')

# Добавляем пустые поля
grouped['Компания'] = ''
grouped['Сфера'] = ''
grouped['Расшифровка ОКПД2'] = ''
grouped['Статус'] = ''

# Упорядочиваем столбцы
grouped = grouped[['Компания', 'Полное наименование', 'ИНН', 'Дата выдачи',
                   'Дата окончания действия', 'Сфера', 'ОКПД2', 'Расшифровка ОКПД2',
                   'Количество российской продукции, шт.', 'Статус']]

grouped.sort_values(by=['ИНН', 'Полное наименование'], inplace=True)

# === Читаем соответствия ===
def clean_string(s):
    if pd.isna(s):
        return s
    return str(s).replace('\xa0', ' ').strip().lower()

# --- Компания ---
company_df = pd.read_excel(mapping_file, sheet_name="Компания_Наименование", engine="openpyxl")
company_df['Полное наименование_norm'] = company_df['Полное наименование'].apply(clean_string)
company_map = pd.Series(company_df['Компания'].values, index=company_df['Полное наименование_norm']).to_dict()
grouped['Полное наименование_norm'] = grouped['Полное наименование'].apply(clean_string)
grouped['Компания'] = grouped['Полное наименование_norm'].map(company_map).fillna('')

# --- Сфера ---
sphere_df = pd.read_excel(mapping_file, sheet_name="Сфера_ОКПД2", engine="openpyxl")
sphere_df['ОКПД2_norm'] = sphere_df['ОКПД2'].astype(str).apply(clean_string)
sphere_map = pd.Series(sphere_df['Сфера'].values, index=sphere_df['ОКПД2_norm']).to_dict()
grouped['ОКПД2_norm'] = grouped['ОКПД2'].astype(str).apply(clean_string)
grouped['Сфера'] = grouped['ОКПД2_norm'].map(sphere_map).fillna('')

# --- Статус ---
status_df = pd.read_excel(mapping_file, sheet_name="Наименование_Статус", engine="openpyxl")
status_df['Полное наименование_norm'] = status_df['Полное наименование'].apply(clean_string)
status_map = pd.Series(status_df['Статус'].values, index=status_df['Полное наименование_norm']).to_dict()
grouped['Статус'] = grouped['Полное наименование_norm'].map(status_map).fillna('')

# Удаляем временные колонки
grouped.drop(columns=['Полное наименование_norm', 'ОКПД2_norm'], inplace=True)

# === Пишем в Excel вручную с форматированием ===
wb = Workbook()
ws = wb.active
ws.title = "Компании"

# Заголовки + строки
for r in dataframe_to_rows(grouped, index=False, header=True):
    ws.append(r)

# Выравнивание
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", horizontal="left")

# === Функция объединения ячеек по колонке ===
def merge_identical_cells(ws, col_letter):
    start_row = 2
    while start_row <= ws.max_row:
        current_value = ws[f"{col_letter}{start_row}"].value
        end_row = start_row
        while end_row + 1 <= ws.max_row and ws[f"{col_letter}{end_row + 1}"].value == current_value:
            end_row += 1
        if end_row > start_row:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
        start_row = end_row + 1

# Объединяем "Полное наименование" (B) и "ИНН" (C)

def merge_status_by_full_name(ws):
    col_status = "J"
    col_name = "B"
    start_row = 2

    while start_row <= ws.max_row:
        name_value = ws[f"{col_name}{start_row}"].value
        status_value = ws[f"{col_status}{start_row}"].value

        end_row = start_row
        while (
            end_row + 1 <= ws.max_row
            and ws[f"{col_name}{end_row + 1}"].value == name_value
            and ws[f"{col_status}{end_row + 1}"].value == status_value
        ):
            end_row += 1

        if end_row > start_row:
            ws.merge_cells(f"{col_status}{start_row}:{col_status}{end_row}")

        start_row = end_row + 1
merge_status_by_full_name(ws)

merge_identical_cells(ws, "A")
merge_identical_cells(ws, "B")
merge_identical_cells(ws, "C")



# Вызов:


# Сохраняем
wb.save(summary_file)
print(f"✅ Всё готово! Файл сохранён с объединёнными ячейками: {summary_file}")
