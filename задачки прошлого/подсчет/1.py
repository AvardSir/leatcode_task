import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Чтение из Excel
df = pd.read_excel("production_res_valid_only copy.xlsx", skiprows=2, engine="openpyxl")

# Переименуем колонки для удобства (если названия могут отличаться — поправь)
df = df.rename(columns={
    'Предприятие': 'Полное наименование',
    'ИНН': 'ИНН',
    'Дата внесения в реестр': 'Дата выдачи',
    'Срок действия': 'Дата окончания действия',
    'ОКПД2': 'ОКПД2'
})

# Оставляем нужные столбцы
df = df[['Полное наименование', 'ИНН', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']].copy()

# Группируем по нужным колонкам и считаем количество повторов ОКПД2
grouped = df.groupby(['ИНН', 'Полное наименование', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']) \
            .size() \
            .reset_index(name='Количество российской продукции, шт.')

# Добавляем пустые столбцы
grouped['Компания'] = ''
grouped['Сфера'] = ''
grouped['Расшифровка ОКПД2'] = ''
grouped['Статус'] = ''

# Переупорядочиваем колонки под итоговый формат
grouped = grouped[['Компания', 'Полное наименование', 'ИНН', 'Дата выдачи',
                   'Дата окончания действия', 'Сфера', 'ОКПД2', 'Расшифровка ОКПД2',
                   'Количество российской продукции, шт.', 'Статус']]

# Сортируем
grouped.sort_values(by=['ИНН', 'Полное наименование'], inplace=True)

# Создаем новую книгу
wb = Workbook()
ws = wb.active
ws.title = "Компании"

# Записываем данные
for r in dataframe_to_rows(grouped, index=False, header=True):
    ws.append(r)

# Выравнивание текста в ячейках
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", horizontal="left")

# Функция для объединения одинаковых ячеек в столбце
def merge_identical_cells(ws, col_letter):
    start_row = 2
    while start_row <= ws.max_row:
        current_value = ws[f"{col_letter}{start_row}"].value
        end_row = start_row
        while end_row + 1 <= ws.max_row and ws[f"{col_letter}{end_row + 1}"].value == current_value:
            end_row += 1
        if end_row > start_row:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
        start_row = end_row + 1

# Объединяем столбцы "Полное наименование" (B) и "ИНН" (C)
merge_identical_cells(ws, "B")
merge_identical_cells(ws, "C")

# Сохраняем файл
output_file = "сводка_по_компаниям.xlsx"
wb.save(output_file)

print(f"✅ Готово! Файл сохранён: {output_file}")
