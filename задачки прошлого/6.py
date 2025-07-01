import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl import Workbook

# Чтение Excel
df = pd.read_excel("production_res_valid_only copy.xlsx", skiprows=2, engine="openpyxl")

# Переименование
df = df.rename(columns={
    'Предприятие': 'Полное наименование',
    'ИНН': 'ИНН',
    'Дата внесения в реестр': 'Дата выдачи',
    'Срок действия': 'Дата окончания действия',
    'ОКПД2': 'ОКПД2'
})

# Только нужные
df = df[['Полное наименование', 'ИНН', 'Дата выдачи', 'Дата окончания действия', 'ОКПД2']].copy()

# Пустые доп. поля
df['Компания'] = ''
df['Сфера'] = ''
df['Расшифровка ОКПД2'] = ''
df['Количество российской продукции, шт.'] = ''
df['Статус'] = ''

# Переупорядочим
df = df[['Компания', 'Полное наименование', 'ИНН', 'Дата выдачи',
         'Дата окончания действия', 'Сфера', 'ОКПД2', 'Расшифровка ОКПД2',
         'Количество российской продукции, шт.', 'Статус']]

# Сортировка для группировки
df.sort_values(by=['Полное наименование', 'ИНН'], inplace=True)

# === Автозаполнение по "Полное наименование" ===
grouped = df.groupby("Полное наименование", sort=False)
for name, group in grouped:
    if len(group) > 1:
        first_row = group.iloc[0]
        for idx in group.index[1:]:
            for col in df.columns:
                if col not in ['ОКПД2', 'Расшифровка ОКПД2', 'Количество российской продукции, шт.']:
                    df.at[idx, col] = first_row[col]

# Создание Excel
wb = Workbook()
ws = wb.active
ws.title = "Компании"

# Запись
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Выравнивание
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center", horizontal="left")

# Объединение одинаковых подряд ячеек
def merge_identical_cells(ws, col_letter):
    start_row = 2
    while start_row <= ws.max_row:
        current_value = ws[f"{col_letter}{start_row}"].value
        end_row = start_row
        while end_row + 1 <= ws.max_row and ws[f"{col_letter}{end_row + 1}"].value == current_value:
            end_row += 1
        if end_row > start_row:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
        start_row = end_row + 1

# Объединяем "Полное наименование" (B) и "ИНН" (C)
merge_identical_cells(ws, "B")
merge_identical_cells(ws, "C")

# Сохраняем
output_file = "сводка_по_компаниям.xlsx"
wb.save(output_file)

print(f"✅ Готово! Файл сохранён: {output_file}")
